#!/usr/bin/env python3
"""
化工管道保温表面积计算程序
支持圆形管道，输出Excel表格
用户可自定义输入参数
"""

import os
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import json

class PipeInsulationCalculator:
    def __init__(self):
        """初始化计算器"""
        # 默认计算参数（可根据需要修改）
        self.default_params = {
            'standard': 'GB/T 4272-2008',  # 采用标准
            'loss_factor': 1.08,           # 施工损耗系数
            'valve_factor': 2.1,           # 阀门面积系数（同直径管道的倍数）
            'elbow_90_factor': 1.5,        # 90°弯头系数
            'elbow_45_factor': 1.2,        # 45°弯头系数
            'tee_factor': 1.8,             # 三通系数
            'reducer_factor': 1.3,         # 变径管系数
            'flange_width': 0.15,          # 法兰保温宽度(m)
            'unit': 'm²',                  # 面积单位
            'precision': 2                 # 小数位数
        }
        
        # 保温材料密度参考（kg/m³）
        self.material_density = {
            '岩棉': 120,
            '玻璃棉': 48,
            '硅酸铝': 200,
            '聚氨酯': 60,
            '橡塑': 95,
            '气凝胶': 180
        }
        
        # 存储用户输入的数据
        self.pipes = []          # 管道列表
        self.fittings = []       # 管件列表
        self.results = []        # 计算结果
        self.project_info = {}   # 项目信息
    
    def get_user_input(self):
        """获取用户输入"""
        print("=" * 60)
        print("化工管道保温表面积计算程序")
        print("=" * 60)
        
        # 获取项目信息
        print("\n【项目信息】")
        self.project_info['project_name'] = input("项目名称: ").strip() or "未命名项目"
        self.project_info['designer'] = input("设计人员: ").strip() or "未指定"
        self.project_info['date'] = datetime.now().strftime('%Y-%m-%d')
        
        # 选择输入方式
        print("\n【输入方式选择】")
        print("1. 手动输入数据")
        print("2. 从JSON文件导入数据")
        print("3. 使用示例数据测试")
        
        choice = input("请选择输入方式 (1-3): ").strip()
        
        if choice == '1':
            self.manual_input()
        elif choice == '2':
            self.import_from_json()
        elif choice == '3':
            self.use_sample_data()
        else:
            print("使用手动输入方式")
            self.manual_input()
    
    def manual_input(self):
        """手动输入数据"""
        print("\n【管道信息输入】")
        print("输入'q'结束管道输入")
        
        pipe_count = 0
        while True:
            pipe_count += 1
            print(f"\n--- 管道 #{pipe_count} ---")
            
            # 检查是否结束
            diameter = input("管道外径(mm) [输入'q'结束]: ").strip()
            if diameter.lower() == 'q':
                if pipe_count == 1:
                    print("至少需要输入一条管道信息")
                    continue
                break
            
            # 验证输入
            try:
                diameter = float(diameter)
                if diameter <= 0:
                    print("错误: 直径必须大于0")
                    continue
            except ValueError:
                print("错误: 请输入有效的数字")
                continue
            
            # 获取其他参数
            length = self.get_float_input("管道长度(m): ", default=1.0)
            quantity = self.get_int_input("管道数量(根): ", default=1)
            insulation_thickness = self.get_float_input("保温层厚度(mm): ", default=50.0)
            material = input("保温材料 [默认: 岩棉]: ").strip() or "岩棉"
            
            # 存储管道信息
            pipe = {
                'id': pipe_count,
                'name': f"管道-{pipe_count}",
                'diameter_mm': diameter,
                'diameter_m': diameter / 1000,  # 转换为米
                'length_m': length,
                'quantity': quantity,
                'insulation_thickness_mm': insulation_thickness,
                'insulation_thickness_m': insulation_thickness / 1000,
                'material': material,
                'type': 'pipe'
            }
            
            self.pipes.append(pipe)
            print(f"已添加: Ø{diameter}mm × {length}m × {quantity}根")
        
        # 输入管件信息
        self.input_fittings()
    
    def get_float_input(self, prompt, default=None):
        """获取浮点数输入"""
        while True:
            value = input(prompt).strip()
            if not value and default is not None:
                return default
            try:
                return float(value)
            except ValueError:
                print("错误: 请输入有效的数字")
    
    def get_int_input(self, prompt, default=None):
        """获取整数输入"""
        while True:
            value = input(prompt).strip()
            if not value and default is not None:
                return default
            try:
                return int(value)
            except ValueError:
                print("错误: 请输入有效的整数")
    
    def input_fittings(self):
        """输入管件信息"""
        print("\n【管件信息输入】")
        print("输入'q'跳过管件输入")
        
        fitting_types = [
            ('阀门', 'valve', '个'),
            ('90°弯头', 'elbow_90', '个'),
            ('45°弯头', 'elbow_45', '个'),
            ('三通', 'tee', '个'),
            ('变径管', 'reducer', '个'),
            ('法兰', 'flange', '对')
        ]
        
        for name, code, unit in fitting_types:
            while True:
                input_str = input(f"{name}数量({unit}) [输入'q'跳过]: ").strip()
                if input_str.lower() == 'q':
                    break
                
                try:
                    quantity = int(input_str)
                    if quantity > 0:
                        # 获取规格
                        if code != 'flange':
                            spec = input(f"{name}规格(DN/mm): ").strip() or "同管道"
                        else:
                            spec = input(f"{name}规格(DN/mm): ").strip() or "DN200"
                        
                        fitting = {
                            'name': name,
                            'type': code,
                            'quantity': quantity,
                            'spec': spec,
                            'unit': unit
                        }
                        self.fittings.append(fitting)
                        print(f"已添加: {name} × {quantity}{unit}")
                    break
                except ValueError:
                    print("错误: 请输入有效的整数")
    
    def import_from_json(self):
        """从JSON文件导入数据"""
        filename = input("请输入JSON文件路径: ").strip()
        if not os.path.exists(filename):
            print(f"文件不存在: {filename}")
            self.manual_input()
            return
        
        try:
            with open(filename, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            if 'pipes' in data:
                self.pipes = data['pipes']
                print(f"已导入 {len(self.pipes)} 条管道信息")
            
            if 'fittings' in data:
                self.fittings = data['fittings']
                print(f"已导入 {len(self.fittings)} 项管件信息")
            
            if 'project_info' in data:
                self.project_info.update(data['project_info'])
            
        except Exception as e:
            print(f"导入失败: {e}")
            self.manual_input()
    
    def use_sample_data(self):
        """使用示例数据"""
        print("使用示例数据进行计算...")
        
        # 示例管道数据
        self.pipes = [
            {
                'id': 1,
                'name': '主蒸汽管道',
                'diameter_mm': 219,
                'diameter_m': 0.219,
                'length_m': 50,
                'quantity': 2,
                'insulation_thickness_mm': 50,
                'insulation_thickness_m': 0.05,
                'material': '岩棉',
                'type': 'pipe'
            },
            {
                'id': 2,
                'name': '循环水管道',
                'diameter_mm': 325,
                'diameter_m': 0.325,
                'length_m': 30,
                'quantity': 1,
                'insulation_thickness_mm': 40,
                'insulation_thickness_m': 0.04,
                'material': '玻璃棉',
                'type': 'pipe'
            }
        ]
        
        # 示例管件数据
        self.fittings = [
            {'name': '阀门', 'type': 'valve', 'quantity': 3, 'spec': 'DN200', 'unit': '个'},
            {'name': '90°弯头', 'type': 'elbow_90', 'quantity': 4, 'spec': 'DN200', 'unit': '个'},
            {'name': '法兰', 'type': 'flange', 'quantity': 6, 'spec': 'DN200', 'unit': '对'}
        ]
        
        # 项目信息
        self.project_info = {
            'project_name': '示例化工厂项目',
            'designer': '张三',
            'date': datetime.now().strftime('%Y-%m-%d')
        }
        
        print("示例数据加载完成")
    
    def calculate_pipe_area(self, pipe):
        """计算单根管道保温面积"""
        # 计算保温后外径
        D = pipe['diameter_m']  # 管道外径(m)
        δ = pipe['insulation_thickness_m']  # 保温厚度(m)
        L = pipe['length_m']  # 管道长度(m)
        
        # 直管段保温面积公式: A = π × (D + 2δ) × L
        area = math.pi * (D + 2 * δ) * L
        
        # 考虑施工损耗
        area_with_loss = area * self.default_params['loss_factor']
        
        return area_with_loss
    
    def calculate_fitting_area(self, fitting, ref_diameter=None):
        """计算管件保温面积"""
        # 如果没有参考直径，使用默认值
        if ref_diameter is None:
            ref_diameter = 0.219  # 默认DN200
        
        # 获取系数
        factor_map = {
            'valve': self.default_params['valve_factor'],
            'elbow_90': self.default_params['elbow_90_factor'],
            'elbow_45': self.default_params['elbow_45_factor'],
            'tee': self.default_params['tee_factor'],
            'reducer': self.default_params['reducer_factor']
        }
        
        if fitting['type'] in factor_map:
            # 阀门、弯头等按同直径管道长度倍数计算
            factor = factor_map[fitting['type']]
            # 假设等效长度为1米
            equivalent_length = 1.0
            area = math.pi * ref_diameter * equivalent_length * factor
        elif fitting['type'] == 'flange':
            # 法兰面积计算
            # 假设法兰外径为管道外径的1.2倍
            flange_diameter = ref_diameter * 1.2
            width = self.default_params['flange_width']
            area = math.pi * flange_diameter * width
        
        # 考虑施工损耗
        area_with_loss = area * self.default_params['loss_factor']
        
        return area_with_loss
    
    def calculate_all(self):
        """计算所有项目的保温面积"""
        print("\n【开始计算】")
        print(f"采用标准: {self.default_params['standard']}")
        print(f"损耗系数: {self.default_params['loss_factor']}")
        
        total_area = 0
        self.results = []
        
        # 计算管道
        for pipe in self.pipes:
            # 单根管道面积
            single_area = self.calculate_pipe_area(pipe)
            # 总面积（考虑数量）
            total_pipe_area = single_area * pipe['quantity']
            
            result = {
                '序号': len(self.results) + 1,
                '名称': pipe['name'],
                '规格': f"Ø{pipe['diameter_mm']}mm",
                '长度/参数': f"{pipe['length_m']}m",
                '数量': pipe['quantity'],
                '保温厚度': f"{pipe['insulation_thickness_mm']}mm",
                '材料': pipe['material'],
                '单件面积(m²)': round(single_area, self.default_params['precision']),
                '总面积(m²)': round(total_pipe_area, self.default_params['precision']),
                '类型': '管道'
            }
            
            self.results.append(result)
            total_area += total_pipe_area
        
        # 计算管件
        for fitting in self.fittings:
            # 需要参考直径，这里使用第一个管道的直径
            ref_diameter = 0.219  # 默认值
            if self.pipes:
                ref_diameter = self.pipes[0]['diameter_m']
            
            # 单件面积
            single_area = self.calculate_fitting_area(fitting, ref_diameter)
            # 总面积
            total_fitting_area = single_area * fitting['quantity']
            
            result = {
                '序号': len(self.results) + 1,
                '名称': fitting['name'],
                '规格': fitting['spec'],
                '长度/参数': '-',
                '数量': fitting['quantity'],
                '保温厚度': '-',
                '材料': '-',
                '单件面积(m²)': round(single_area, self.default_params['precision']),
                '总面积(m²)': round(total_fitting_area, self.default_params['precision']),
                '类型': '管件'
            }
            
            self.results.append(result)
            total_area += total_fitting_area
        
        # 添加汇总行
        summary = {
            '序号': '汇总',
            '名称': '总计',
            '规格': '-',
            '长度/参数': '-',
            '数量': sum(r['数量'] for r in self.results if r['类型'] == '管道'),
            '保温厚度': '-',
            '材料': '-',
            '单件面积(m²)': '-',
            '总面积(m²)': round(total_area, self.default_params['precision']),
            '类型': '汇总'
        }
        
        self.results.append(summary)
        
        print(f"计算完成！总保温面积: {round(total_area, 2)} m²")
        return total_area
    
    def create_excel_report(self, filename=None):
        """创建Excel报告"""
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"管道保温计算_{timestamp}.xlsx"
        
        print(f"\n正在生成Excel报告: {filename}")
        
        # 创建工作簿
        wb = Workbook()
        
        # 1. 封面页
        self.create_cover_page(wb)
        
        # 2. 计算明细页
        self.create_detail_page(wb)
        
        # 3. 材料汇总页
        self.create_summary_page(wb)
        
        # 4. 参数说明页
        self.create_parameter_page(wb)
        
        # 保存文件
        wb.save(filename)
        print(f"Excel报告已保存: {filename}")
        
        return filename
    
    def create_cover_page(self, wb):
        """创建封面页"""
        if 'Sheet' in wb.sheetnames:
            ws = wb['Sheet']
            ws.title = "封面"
        else:
            ws = wb.create_sheet("封面")
        
        # 设置列宽
        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # 标题
        title_font = Font(name='微软雅黑', size=20, bold=True)
        ws['A1'] = "化工管道保温表面积计算报告"
        ws['A1'].font = title_font
        ws.merge_cells('A1:J1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 项目信息
        ws['A3'] = "项目名称:"
        ws['B3'] = self.project_info.get('project_name', '未命名项目')
        ws.merge_cells('B3:J3')
        
        ws['A4'] = "设计人员:"
        ws['B4'] = self.project_info.get('designer', '未指定')
        ws.merge_cells('B4:J4')
        
        ws['A5'] = "计算日期:"
        ws['B5'] = self.project_info.get('date', datetime.now().strftime('%Y-%m-%d'))
        ws.merge_cells('B5:J5')
        
        ws['A6'] = "采用标准:"
        ws['B6'] = self.default_params['standard']
        ws.merge_cells('B6:J6')
        
        # 计算参数

        ws['A8'] = "计算参数"
        ws['A8'].font = Font(bold=True)
        ws.merge_cells('A8:J8')
        
        row = 9
        params = [
            ("施工损耗系数", f"{self.default_params['loss_factor']}"),
            ("阀门面积系数", f"{self.default_params['valve_factor']}"),
            ("90°弯头系数", f"{self.default_params['elbow_90_factor']}"),
            ("45°弯头系数", f"{self.default_params['elbow_45_factor']}"),
            ("三通系数", f"{self.default_params['tee_factor']}"),
            ("变径管系数", f"{self.default_params['reducer_factor']}"),
            ("法兰保温宽度", f"{self.default_params['flange_width']} m"),
            ("面积单位", self.default_params['unit'])
        ]
        
        for name, value in params:
            ws[f'A{row}'] = name
            ws[f'B{row}'] = value
            ws.merge_cells(f'B{row}:J{row}')
            row += 1
        
        # 添加边框

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=row-1, min_col=1, max_col=10):
            for cell in row:
                cell.border = thin_border
    
    def create_detail_page(self, wb):
        """创建计算明细页"""
        ws = wb.create_sheet("计算明细")
        
        # 设置列宽

        column_widths = [8, 20, 15, 15, 8, 12, 15, 15, 15, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 标题

        title_font = Font(name='微软雅黑', size=14, bold=True)
        ws['A1'] = "管道保温面积计算明细表"
        ws['A1'].font = title_font
        ws.merge_cells('A1:J1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 表头

        headers = ['序号', '名称', '规格', '长度/参数', '数量', '保温厚度', '材料', 
                  '单件面积(m²)', '总面积(m²)', '类型']
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 数据行

        row = 4
        for result in self.results:
            for col, key in enumerate(headers, 1):
                value = result.get(key, '')
                cell = ws.cell(row=row, column=col, value=value)
                
                # 汇总行特殊格式

                if result['类型'] == '汇总':
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
                # 数值对齐

                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='center')
            
            row += 1
        
        # 添加边框

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_cells in ws.iter_rows(min_row=3, max_row=row-1, min_col=1, max_col=10):
            for cell in row_cells:
                cell.border = thin_border
    
    def create_summary_page(self, wb):
        """创建材料汇总页"""

        ws = wb.create_sheet("材料汇总")
        
        # 按材料汇总

        material_summary = {}
        for result in self.results:
            if result['类型'] == '管道':
                material = result['材料']
                area = result['总面积(m²)']
                if isinstance(area, (int, float)):
                    if material not in material_summary:
                        material_summary[material] = 0
                    material_summary[material] += area
        
        # 标题

        ws['A1'] = "保温材料用量汇总"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:C1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 表头

        headers = ['保温材料', '总面积(m²)', '估算用量(kg)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
        
        # 数据

        row = 4
        total_area = 0
        total_weight = 0
        
        for material, area in material_summary.items():
            # 估算重量（面积 × 假设厚度0.05m × 材料密度）

            density = self.material_density.get(material, 120)  # 默认120kg/m³
            thickness = 0.05  # 假设平均厚度5cm
            weight = area * thickness * density
            
            ws.cell(row=row, column=1, value=material)
            ws.cell(row=row, column=2, value=round(area, 2))
            ws.cell(row=row, column=3, value=round(weight, 1))
            
            total_area += area
            total_weight += weight
            row += 1
        
        # 总计行

        ws.cell(row=row, column=1, value="总计").font = Font(bold=True)
        ws.cell(row=row, column=2, value=round(total_area, 2)).font = Font(bold=True)
        ws.cell(row=row, column=3, value=round(total_weight, 1)).font = Font(bold=True)
    
    def create_parameter_page(self, wb):
        """创建参数说明页"""

        ws = wb.create_sheet("参数说明")
        
        ws['A1'] = "计算公式和参数说明"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:C1')
        
        # 计算公式

        formulas = [
            ("1. 直管段保温面积", "A = π × (D + 2δ) × L"),
            ("  其中:", "D - 管道外径(m)"),
            ("  ", "δ - 保温层厚度(m)"),
            ("  ", "L - 管道长度(m)"),
            ("", ""),
            ("2. 阀门保温面积", "A = 阀门系数 × π × D × Lₑ"),
            ("  其中:", "Lₑ - 等效长度(通常取1m)"),
            ("", ""),
            ("3. 弯头保温面积", "A = 弯头系数 × π × D × Lₑ"),
            ("  90°弯头系数:", f"{self.default_params['elbow_90_factor']}"),
            ("  45°弯头系数:", f"{self.default_params['elbow_45_factor']}"),
            ("", ""),
            ("4. 施工损耗", "最终面积 = 计算面积 × 损耗系数"),
            ("  损耗系数:", f"{self.default_params['loss_factor']}")
        ]
        
        row = 3
        for left, right in formulas:
            ws.cell(row=row, column=1, value=left)
            ws.cell(row=row, column=2, value=right)
            row += 1
        
        # 设置列宽

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
    
    def export_to_json(self, filename=None):
        """导出数据到JSON文件，用于后续修改"""

        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"管道数据_{timestamp}.json"
        
        data = {
            'project_info': self.project_info,
            'pipes': self.pipes,
            'fittings': self.fittings,
            'calculation_params': self.default_params,
            'results': self.results
        }
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        print(f"数据已导出到: {filename}")
        return filename

def main():
    """主函数"""
    print("=" * 60)
    print("化工管道保温表面积计算程序")
    print("=" * 60)
    print("功能说明:")
    print("1. 支持圆形管道保温面积计算")
    print("2. 支持阀门、弯头、法兰等管件计算")
    print("3. 输出详细的Excel计算报告")
    print("4. 可保存和加载计算数据")
    print("=" * 60)
    
    # 创建计算器实例

    calculator = PipeInsulationCalculator()
    
    # 获取用户输入

    calculator.get_user_input()
    
    # 进行计算

    total_area = calculator.calculate_all()
    
    # 生成Excel报告

    excel_file = calculator.create_excel_report()
    
    # 导出数据（可选）

    export_choice = input("\n是否导出数据到JSON文件以便后续修改? (y/n): ").lower()
    if export_choice == 'y':
        json_file = calculator.export_to_json()
        print(f"数据文件: {json_file}")
    
    print("\n" + "=" * 60)
    print("计算完成！")
    print(f"总保温面积: {total_area:.2f} m²")
    print(f"Excel报告: {excel_file}")
    print("=" * 60)
    
    # 打开Excel文件（如果系统支持）

    open_excel = input("\n是否打开Excel文件? (y/n): ").lower()
    if open_excel == 'y':
        try:
            os.startfile(excel_file)  # Windows
        except:
            print(f"请手动打开文件: {excel_file}")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n程序已中断")
    except Exception as e:
        print(f"程序运行出错: {e}")
        print("请确保已安装所需库: pip install openpyxl")
    finally:
        input("\n按回车键退出...")
start_pipe_calc.bat（启动脚本）：

@echo off
chcp 65001 >nul
title 化工管道保温表面积计算程序

echo ================================================
echo       化工管道保温表面积计算程序
echo ================================================
echo.

REM 检查Python是否安装
echo 检查Python环境...
python --version >nul 2>&1
if errorlevel 1 (
    echo ❌ 未找到Python！
    echo.
    echo 请先安装Python 3.x：
    echo 1. 访问 https://www.python.org/downloads/
    echo 2. 下载并安装Python
    echo 3. 安装时务必勾选 "Add Python to PATH"
    echo.
    pause
    exit /b 1
)

REM 检查openpyxl库
echo 检查所需库...
python -c "import openpyxl" >nul 2>&1
if errorlevel 1 (
    echo ⚠ 缺少openpyxl库，正在安装...
    pip install openpyxl
    if errorlevel 1 (
        echo ❌ 安装失败，请手动安装：
        echo pip install openpyxl
        pause
        exit
